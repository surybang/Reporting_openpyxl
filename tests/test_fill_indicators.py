import pytest

from openpyxl import Workbook, load_workbook

from reporting_generator.fill_indicators import fill_indicators


@pytest.fixture
def minimal_workbook(tmp_path):
    """Create a minimal workbook with 2 sheets (DATA, Indicateurs)"""
    path = tmp_path / "wb.xlsx"
    wb = Workbook()
    wb.active.title = "DATA"
    wb.create_sheet("Indicateurs")
    wb.save(path)
    wb.close()
    return str(path)


def test_fill_indicators_integration(minimal_workbook):
    """Tests values for E8, E9, E10"""
    file_path = minimal_workbook

    fill_indicators(file_path, data_sheet="DATA")

    wb = load_workbook(file_path, data_only=False)
    ws = wb["Indicateurs"]

    # Vérifions quelques formules
    assert ws["E8"].value == '=COUNTIF(DATA!B:B, "PP")'
    assert ws["E9"].value == '=COUNTIF(DATA!B:B, "PM")'
    assert ws["E10"].value == "=SUM(E8:E9)"

    wb.close()
