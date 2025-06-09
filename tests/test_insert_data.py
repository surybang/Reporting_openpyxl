import pandas as pd
from openpyxl import load_workbook

from reporting_generator.insert_data import write_data_to_excel


def test_write_data_to_excel(tmp_path):
    # Setup : créer un DataFrame factice
    df = pd.DataFrame({"Nom": ["John", "Doe"], "Âge": [30, 25]})

    # Créer un chemin vers un fichier temporaire
    path_file = tmp_path / "test.xlsx"

    # Créer un fichier Excel vide (obligatoire pour 'mode=a')
    with pd.ExcelWriter(path_file, engine="openpyxl") as writer:
        writer.book.create_sheet("dummy")

    # Appel de la fonction
    write_data_to_excel(str(path_file), df, sheet_name="DATA")

    # Vérifier que les données ont été écrites
    wb = load_workbook(filename=path_file)
    ws = wb["DATA"]

    assert ws["A1"].value == "Nom"
    assert ws["B1"].value == "Âge"
    assert ws["A2"].value == "John"
    assert ws["B2"].value == 30
    assert ws["A3"].value == "Doe"
    assert ws["B3"].value == 25
