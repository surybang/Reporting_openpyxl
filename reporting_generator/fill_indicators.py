from openpyxl import load_workbook


def fill_indicators(path_file: str, data_sheet: str = 'DATA') -> None:
    """
    Fill the indicators in the Excel file with formulas.

    Args:
        path_file (str): Path to the Excel file.
        data_sheet (str, optional): Name of the data sheet. Defaults to 'DATA'.
    """
    wb = load_workbook(path_file)
    ws = wb['Indicateurs']

    def formula_COUNTIF(col, val):
        return f'=COUNTIF({data_sheet}!{col}:{col}, "{val}")'

    def formula_COUNTIFS(pairs):
        conditions = ', '.join(f'{data_sheet}!{col}:{col}, "{val}"' for col, val in pairs)
        return f'=COUNTIFS({conditions})'

    def formula_SUM(range_str):
        return f'=SUM({range_str})'

    def formula_CALCUL(expr):
        return f'={expr}'

    config = [
        # Répartition PP/PM
        {'row': 8, 'formule': 'COUNTIF', 'args': [('B', 'PP')]},
        {'row': 9, 'formule': 'COUNTIF', 'args': [('B', 'PM')]},
        {'row': 10, 'formule': 'SUM', 'args': ['E8:E9']},

        # Scores V/O/R
        {'row': 14, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'V')]},
        {'row': 15, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'O')]},
        {'row': 16, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'R')]},
        {'row': 17, 'formule': 'CALCUL', 'args': 'E10 - (E14+E15+E16)'},
        {'row': 18, 'formule': 'SUM', 'args': ['E14+E15+E16']},

        # DRC Complet
        {'row': 22, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('G', "VRAI")]},
        {'row': 23, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('G', "VRAI")]},
        {'row': 24, 'formule': 'SUM', 'args': ['E22:E23']},

        # Focus V/O
        {'row': 28, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'V')]},
        {'row': 29, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'V')]},
        {'row': 30, 'formule': 'SUM', 'args': ['E28:E29']},

        {'row': 31, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'O')]},
        {'row': 32, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'O')]},
        {'row': 33, 'formule': 'SUM', 'args': ['E31:E32']},

        # Focus V/O avec DRC complet
        {'row': 34, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'V'), ('G', 'VRAI')]},
        {'row': 35, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'V'), ('G', 'VRAI')]},
        {'row': 36, 'formule': 'SUM', 'args': ['E34:E35']},
        {'row': 37, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'O'), ('G', 'VRAI')]},
        {'row': 38, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'O'), ('G', 'VRAI')]},
        {'row': 39, 'formule': 'SUM', 'args': ['E37:E38']},

        # Focus sur les R
        {'row': 43, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'R')]},
        {'row': 44, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'R')]},
        {'row': 45, 'formule': 'SUM', 'args': ['E43:E44']},

        # Focus sur les R scorés automatiquement ou par un agent
        {'row': 46, 'formule': 'COUNTIFS', 'args': [('D', 'R'), ('F', 'AUTO')]},
        {'row': 47, 'formule': 'COUNTIFS', 'args': [('D', 'R'), ('F', '<> AUTO')]},
        {'row': 48, 'formule': 'SUM', 'args': ['E46:E47']},

        # Focus sur les scorés R avec DRC complet
        {'row': 49, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('D', 'R'), ('G', 'VRAI')]},
        {'row': 50, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('D', 'R'), ('G', 'VRAI')]},
        {'row': 51, 'formule': 'SUM', 'args': ['E49:E50']},

        # Focus sur les nouveaux clients (Z score prev)
        {'row': 52, 'formule': 'COUNTIFS', 'args': [('B', 'PP'), ('E', 'Z'), ('G', 'VRAI')]},
        {'row': 53, 'formule': 'COUNTIFS', 'args': [('B', 'PM'), ('E', 'Z'), ('G', 'VRAI')]},
        {'row': 54, 'formule': 'SUM', 'args': ['E52:E53']},
    ]

    for item in config:
        row = item['row']
        cell = f'E{row}'
        formule = item['formule']
        args = item['args']

        if formule == 'COUNTIF':
            ws[cell] = formula_COUNTIF(*args[0])
        elif formule == 'COUNTIFS':
            ws[cell] = formula_COUNTIFS(args)
        elif formule == 'SUM':
            ws[cell] = formula_SUM(args[0])
        elif formule == 'CALCUL':
            ws[cell] = formula_CALCUL(args)
        else:
            raise ValueError(f"Unknown formula : {formule}")

    wb.save(path_file)
    wb.close()
